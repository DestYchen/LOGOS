#!/usr/bin/env python
"""Convert each visible sheet in an Excel file into a separate PDF via LibreOffice.

Variant 1: keep formulas, export via LibreOffice; for per-sheet PDF we hide other sheets
in a temporary copy.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl


def safe_filename(name: str) -> str:
    # Replace illegal filename chars
    safe = re.sub(r"[\\/:*?\"<>|]", "_", name.strip())
    safe = re.sub(r"\s+", "_", safe)
    return safe or "sheet"


def find_soffice() -> str | None:
    # Allow overriding via environment variable
    env_path = os.environ.get("SOFFICE_PATH")
    if env_path:
        candidate = Path(env_path)
        if candidate.exists():
            return str(candidate)

    for candidate in ("soffice", "soffice.exe"):
        path = shutil.which(candidate)
        if path:
            return path

    # Fallback: query Windows 'where'
    try:
        result = subprocess.run(
            ["where", "soffice"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        for line in result.stdout.splitlines():
            line = line.strip()
            if line.lower().endswith(("soffice.exe", "soffice.com")) and Path(line).exists():
                return line
    except Exception:
        pass

    return None


def convert_sheet(excel_path: Path, sheet_name: str, output_dir: Path, soffice: str) -> Path:
    temp_path = output_dir / f"__temp__{safe_filename(sheet_name)}.xlsx"
    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet not found: {sheet_name}")

    # Move target sheet to the first position and make it active.
    sheet = wb[sheet_name]
    current_index = wb.sheetnames.index(sheet_name)
    if current_index != 0:
        wb.move_sheet(sheet, offset=-current_index)
    wb.active = 0

    for name in wb.sheetnames:
        ws = wb[name]
        is_target = name == sheet_name
        ws.sheet_state = "visible" if is_target else "hidden"
        ws.sheet_view.tabSelected = is_target
        if not is_target:
            # Clear print ranges on other sheets; LibreOffice prints sheets with print areas
            # even when they are hidden.
            ws.print_area = None
    wb.save(temp_path)
    wb.close()

    # Convert to PDF
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(output_dir), str(temp_path)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    temp_path.unlink(missing_ok=True)

    pdf_path = output_dir / f"{temp_path.stem}.pdf"
    final_path = output_dir / f"{safe_filename(sheet_name)}.pdf"
    if pdf_path.exists():
        pdf_path.replace(final_path)
    return final_path


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: excel_to_pdf_per_sheet.py <path-to-xlsx>")
        return 1

    excel_path = Path(sys.argv[1]).expanduser().resolve()
    if not excel_path.exists():
        print(f"File not found: {excel_path}")
        return 1

    soffice = find_soffice()
    if not soffice:
        print("LibreOffice (soffice) not found in PATH.")
        return 1

    output_dir = excel_path.parent / "sheet_pdfs"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
    sheet_names = [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
    wb.close()

    print(f"Converting {len(sheet_names)} sheets from {excel_path.name}...")
    for name in sheet_names:
        pdf_path = convert_sheet(excel_path, name, output_dir, soffice)
        print(f"- {name} -> {pdf_path.name}")

    print(f"Done. Output: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
