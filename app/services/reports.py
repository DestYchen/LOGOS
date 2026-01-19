from __future__ import annotations

import io
import json
import uuid
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.core.storage import batch_dir


def report_path(batch_id: uuid.UUID) -> Path:
    return batch_dir(str(batch_id)).report / "report.json"


def load_report(batch_id: uuid.UUID) -> Dict[str, Any]:
    path = report_path(batch_id)
    if not path.exists():
        raise FileNotFoundError
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _document_rows(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for document in report.get("documents", []):
        base = {
            "doc_id": document.get("doc_id", ""),
            "filename": document.get("filename", ""),
            "doc_type": document.get("doc_type", ""),
            "status": document.get("status", ""),
        }
        fields = document.get("fields", {}) or {}
        if fields:
            for field_key, field_payload in fields.items():
                payload = field_payload or {}
                rows.append(
                    {
                        **base,
                        "field_key": field_key,
                        "value": payload.get("value"),
                        "confidence": payload.get("confidence"),
                        "source": payload.get("source"),
                        "page": payload.get("page"),
                    }
                )
        else:
            rows.append({**base, "field_key": "", "value": None, "confidence": None, "source": None, "page": None})
    return rows


def _extract_document_matrix(report: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    for item in report.get("validations", []):
        if item.get("rule_id") == "document_matrix":
            refs = item.get("refs") or []
            if refs and isinstance(refs[0], dict):
                matrix = refs[0]
                documents = matrix.get("documents")
                rows = matrix.get("rows")
                if isinstance(documents, list) and isinstance(rows, list):
                    return matrix
    return None


def _extract_document_matrix_diff(report: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    for item in report.get("validations", []):
        if item.get("rule_id") == "document_matrix_diff":
            refs = item.get("refs") or []
            if refs and isinstance(refs[0], dict):
                matrix = refs[0]
                documents = matrix.get("documents")
                rows = matrix.get("rows")
                if isinstance(documents, list) and isinstance(rows, list):
                    return matrix
    return None


def _summarize_severity(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        sev = row.get("severity") or ""
        counts[sev] = counts.get(sev, 0) + 1
    return counts


def _validation_rows(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for item in report.get("validations", []):
        if item.get("rule_id") in ("document_matrix", "document_matrix_diff"):
            continue
        refs = item.get("refs") or []
        formatted_refs = _format_references(refs)
        rows.append(
            {
                "rule_id": item.get("rule_id", ""),
                "severity": item.get("severity", ""),
                "message": item.get("message", ""),
                "refs": formatted_refs,
            }
        )
    return rows


def _format_references(refs: List[Any]) -> Dict[str, Any]:
    if not refs:
        return {}
    # Flatten dicts so that each document contributes field/value pairs
    doc_rows: Dict[str, Dict[str, str]] = {}
    for ref in refs:
        if not isinstance(ref, dict):
            continue
        doc_id = ref.get("doc_id") or ""
        key = ref.get("field_key") or ""
        value = ref.get("value")
        if key:
            doc_rows.setdefault(doc_id, {})[key] = value if value is not None else ""
    return doc_rows


def build_report_tables(report: Dict[str, Any]) -> Tuple[
    Optional[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]
]:
    matrix = _extract_document_matrix(report)
    return matrix, _document_rows(report), _validation_rows(report)


def extract_document_matrix_diff(report: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    return _extract_document_matrix_diff(report)


def _write_table(sheet, headers: List[str], rows: List[List[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def export_report_excel(report: Dict[str, Any]) -> io.BytesIO:
    field_matrix, document_rows, validation_rows = build_report_tables(report)

    wb = Workbook()
    ws_matrix = wb.active
    ws_matrix.title = "Field Matrix"
    status_colors = {
        "anchor": PatternFill("solid", fgColor="BDD7EE"),
        "match": PatternFill("solid", fgColor="C6E0B4"),
        "missing": PatternFill("solid", fgColor="F8CBAD"),
        "mismatch": PatternFill("solid", fgColor="F4B084"),
    }
    if field_matrix:
        doc_headers = ["FieldKey"] + field_matrix.get("documents", [])
        ws_matrix.append(doc_headers)
        for row in field_matrix.get("rows", []):
            statuses = row.get("statuses", {}) or {}
            values = [row.get(h, "") for h in doc_headers]
            ws_matrix.append(values)
            current_row = ws_matrix.max_row
            for idx, header in enumerate(doc_headers, start=1):
                status = statuses.get(header)
                if status and header != "FieldKey":
                    cell = ws_matrix.cell(row=current_row, column=idx)
                    fill = status_colors.get(status)
                    if fill:
                        cell.fill = fill
    else:
        ws_matrix.append(["FieldKey"])

    ws_docs = wb.create_sheet("Documents")
    doc_headers = ["Document ID", "Filename", "Type", "Status", "Field", "Value", "Confidence", "Source", "Page"]
    doc_values = [
        [
            row.get("doc_id"),
            row.get("filename"),
            row.get("doc_type"),
            row.get("status"),
            row.get("field_key"),
            row.get("value"),
            row.get("confidence"),
            row.get("source"),
            row.get("page"),
        ]
        for row in document_rows
    ]
    _write_table(ws_docs, doc_headers, doc_values)

    raw_validations = [
        item
        for item in report.get("validations", [])
        if item.get("rule_id") not in ("document_matrix", "document_matrix_diff")
    ]
    ws_validations = wb.create_sheet("Validations")
    ws_validations.append(["Message", "Reference"])
    for item in raw_validations:
        message = _translate_validation_message(item.get("message", ""))
        headers, values = _flatten_reference_rows(item.get("refs") or [])
        if not headers:
            headers = ["Отсутствует"]
            values = ["Отсутствует"]
        ws_validations.append([message] + headers)
        ws_validations.append([""] + values)

    _write_products_sheet(wb, report)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_report_excel_for_batch(batch_id: uuid.UUID) -> io.BytesIO:
    report = load_report(batch_id)
    return export_report_excel(report)


def _translate_validation_message(message: str) -> str:
    translations = {
        "missing or invalid inputs for date comparison": "пропущены даты или значения невалидны",
        "missing or invalid inputs for comparison": "пропущены данные или значения невалидны",
        "missing or invalid anchor value": "значение опорного поля отсутствует или неверно",
        "values are not equal across documents": "значения отличаются между документами",
    }
    result = message or ""
    for eng, rus in translations.items():
        result = result.replace(eng, rus)
    return result


def _flatten_reference_rows(refs: List[Any]) -> Tuple[List[str], List[str]]:
    headers: List[str] = []
    values: List[str] = []
    seen: Dict[str, int] = {}
    for ref in refs:
        if not isinstance(ref, dict):
            continue
        field_key = ref.get("field_key")
        if not field_key:
            continue
        doc_label = ref.get("doc_type") or ref.get("doc_id")
        header = field_key
        if field_key in seen:
            seen[field_key] += 1
            suffix = doc_label or seen[field_key]
            header = f"{field_key} ({suffix})"
        else:
            seen[field_key] = 1
        value = ref.get("value")
        if ref.get("present") is False or value in (None, ""):
            display = "Отсутствует"
        else:
            display = str(value)
        headers.append(header)
        values.append(display)
    return headers, values


def _format_product_entry(entry: Any) -> str:
    if isinstance(entry, dict):
        value = entry.get("value")
        confidence = entry.get("confidence")
        if value in (None, ""):
            value_str = "Отсутствует"
        else:
            value_str = str(value)
        if confidence not in (None, ""):
            try:
                conf_val = float(confidence)
                return f"{value_str}\n({conf_val:.2f})"
            except (ValueError, TypeError):
                return f"{value_str}\n({confidence})"
        return value_str
    if entry in (None, ""):
        return "Отсутствует"
    return str(entry)


def _write_products_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    comparisons = report.get("product_comparisons") or []
    if not comparisons:
        return
    ws = wb.create_sheet("Products")
    attribute_rows = [
        ("Наименование", "name_product"),
        ("Латинское название", "latin_name"),
        ("Размер", "size_product"),
        ("Единица / коробка", "unit_box"),
        ("Кол-во упаковок", "packages"),
        ("Цена за единицу", "price_per_unit"),
        ("Сумма", "total_price"),
    ]
    for idx, comparison in enumerate(comparisons, start=1):
        docs = comparison.get("documents") or []
        if not docs:
            continue
        product_title = comparison.get("product_key", {}).get("name_product") or f"Product #{idx}"
        ws.append([product_title])
        ws.append([""] + [f"{doc.get('doc_type', '')} ({doc.get('product_id', '')})" for doc in docs])
        for label, field_key in attribute_rows:
            row = [label]
            for doc in docs:
                entry = (doc.get("fields") or {}).get(field_key)
                row.append(_format_product_entry(entry))
            ws.append(row)
        ws.append([])
